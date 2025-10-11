import re
import time
import random
import requests
from bs4 import BeautifulSoup
import openpyxl
from requests.exceptions import RequestException



SKILL_BADGES = [
    "The Basics of Google Cloud Compute [Skill Badge]",
    "Get Started with Cloud Storage [Skill Badge]",
    "Get Started with Pub/Sub [Skill Badge]",
    "Get Started with API Gateway [Skill Badge]",
    "Get Started with Looker [Skill Badge]",
    "Get Started with Dataplex [Skill Badge]",
    "Get Started with Google Workspace Tools [Skill Badge]",
    "App Building with AppSheet [Skill Badge]",
    "Develop with Apps Script and AppSheet [Skill Badge]",
    "Build a Website on Google Cloud [Skill Badge]",
    "Set Up a Google Cloud Network [Skill Badge]",
    "Store, Process, and Manage Data on Google Cloud - Console [Skill Badge]",
    "Cloud Run Functions: 3 Ways [Skill Badge]",
    "App Engine: 3 Ways [Skill Badge]",
    "Cloud Speech API: 3 Ways [Skill Badge]",
    "Monitoring in Google Cloud [Skill Badge]",
    "Analyze Speech and Language with Google APIs [Skill Badge]",
    "Prompt Design in Vertex AI [Skill Badge]",
    "Develop Gen AI Apps with Gemini and Streamlit [Skill Badge]"
]

ARCADE = [
    "Level 3: Google Cloud Adventures",
    "Diwali in The Arcade",
    "Level 3: Generative AI [Game]",
    "Level 2: Generative AI [Game]",
    "Level 1: Generative AI [Game]"
]

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
DEBUG = False  


def normalize(s: str) -> str:
    """Normalize text for fuzzy matching."""
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r'[:\-\–\—/]', ' ', s)
    s = re.sub(r'[^a-z0-9 ]+', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

norm_skill_map = {normalize(x): x for x in SKILL_BADGES}
norm_arcade_map = {normalize(x): x for x in ARCADE}


wb = openpyxl.load_workbook('B.xlsx')
sheet = wb.active
max_row = sheet.max_row


if sheet.cell(row=1, column=7).value is None:
    sheet.cell(row=1, column=7).value = "# of Skill Badges"
if sheet.cell(row=1, column=8).value is None:
    sheet.cell(row=1, column=8).value = "Names of Skill Badges"
if sheet.cell(row=1, column=9).value is None:
    sheet.cell(row=1, column=9).value = "# of Arcade Games"
if sheet.cell(row=1, column=10).value is None:
    sheet.cell(row=1, column=10).value = "Names of Completed Arcade Games"
if sheet.cell(row=1, column=11).value is None:
    sheet.cell(row=1, column=11).value = "Eligible for Goodies"

# --- MAIN LOOP ---
for i in range(2, max_row + 1):  # skip header
    url = sheet.cell(row=i, column=3).value
    if not url:
        print(f"Row {i}: No URL found, skipping.")
        continue

    print(f"\nFetching data for Row {i} → {url}")

    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
    except RequestException as e:
        print(f"Error fetching row {i}: {e}")
        sheet.cell(row=i, column=7).value = "Error"
        sheet.cell(row=i, column=9).value = "Error"
        sheet.cell(row=i, column=11).value = "Error"
        continue

    soup = BeautifulSoup(resp.content, "html.parser")

    results = soup.find_all("span", class_="ql-title-medium l-mts")

    found_skills = set()
    found_arcades = set()

    if DEBUG:
        print("  Found badges:")

    for r in results:
        raw = r.get_text(" ", strip=True)
        n = normalize(raw)
        if DEBUG:
            print("   -", raw)

        # Skill badge detection
        if n in norm_skill_map:
            found_skills.add(norm_skill_map[n])
        else:
            for k, orig in norm_skill_map.items():
                if k and (k in n or n in k):
                    found_skills.add(orig)

        # Arcade detection
        if n in norm_arcade_map:
            found_arcades.add(norm_arcade_map[n])
        else:
            for k, orig in norm_arcade_map.items():
                if k and (k in n or n in k):
                    found_arcades.add(orig)

        if "arcade" in n or "game" in n:
            found_arcades.add(raw)

    
    num_skills = len(found_skills)
    num_arcades = len(found_arcades)

    
    eligible = "TRUE" if num_skills == 19 and num_arcades >= 1 else "FALSE"

    
    sheet.cell(row=i, column=7).value = num_skills
    sheet.cell(row=i, column=8).value = ", ".join(sorted(found_skills))
    sheet.cell(row=i, column=9).value = num_arcades
    sheet.cell(row=i, column=10).value = ", ".join(sorted(found_arcades))
    sheet.cell(row=i, column=11).value = eligible

    print(f"→ Skill Badges: {num_skills} | Arcade Games: {num_arcades} | Eligible: {eligible}")

    
    time.sleep(random.uniform(1.5, 3.0))


wb.save("H.xlsx")
print("\n✅ All participant data updated successfully! Saved to 'H.xlsx'")
